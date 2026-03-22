import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Property Grid"

DARK = "0C0A18"
ORANGE = "E86A2D"
DARK_LIGHT = "1E1A30"

thin = Border(
    left=Side(style='thin', color='2A2640'),
    right=Side(style='thin', color='2A2640'),
    top=Side(style='thin', color='2A2640'),
    bottom=Side(style='thin', color='2A2640')
)

header_fill = PatternFill('solid', fgColor=DARK)
header_font = Font(name='Arial', size=9, bold=True, color=ORANGE)
starknet_fill = PatternFill('solid', fgColor=ORANGE)
starknet_font = Font(name='Arial', size=10, bold=True, color=DARK)
name_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
row_fill_even = PatternFill('solid', fgColor='151225')
row_fill_odd = PatternFill('solid', fgColor=DARK_LIGHT)
tick_font = Font(name='Arial', size=13, bold=True, color=ORANGE)
partial_font = Font(name='Arial', size=11, bold=True, color='C4883D')  # dimmer orange for partial
dash_font = Font(name='Arial', size=10, color='3D3658')
title_font = Font(name='Arial', size=16, bold=True, color=ORANGE)
subtitle_font = Font(name='Arial', size=9, italic=True, color='7A7490')
center = Alignment(horizontal='center', vertical='center', wrap_text=True)

features = [
    # Privacy model
    "Multi-asset\nprivacy pool",
    "Formally ZK\nproofs",
    "Secretless\n(no per-note keys)",
    "Efficient note\ndiscovery",
    "Selective\ndisclosure",
    # Cryptographic foundation
    "No trusted\nsetup",
    "Post-quantum\nproof system",
    "Succinct\nverifiability",
    # Performance & UX
    "Fast shield/\nunshield",
    "End-to-end\ntransaction time",
    "Scalable private\ntransactions",
    "Programmable\nprivacy",
    # Ecosystem & access
    "DeFi\ncomposability",
    "Existing\necosystem",
    "EVM\ncompatible",
    "Account\nabstraction",
    "Hardware wallet\nsupport",
]

short_defs = [
    "one pool hides\nall token types",
    "ZK property\nformally proven",
    "one key, no\nsecrets to lose",
    "find your\nfunds fast",
    "reveal only\nwhat you choose",
    "no ceremony,\nno trust needed",
    "safe from\nquantum computers",
    "anyone can\nverify the chain",
    "go private or\npublic in seconds",
    "proof +\nconfirmation speed",
    "high throughput,\nlow cost",
    "complex DeFi ops,\nfully private",
    "works with\nexisting protocols",
    "real users,\nreal liquidity",
    "runs on\nEVM chains",
    "multisig, recovery,\nsession keys",
    "sign with\nLedger/Trezor",
]

# Tick values: True = full tick, "partial" = (✓) with footnote, False = dash

footnotes = {
    1: "incoming transactions only. can\u2019t reliably view outgoing or verify balance. all-or-nothing, no per-transaction selectivity",
    2: "Orchard pool only (since May 2022). Sapling pool still relies on a trusted setup ceremony",
    3: "composable private/public functions exist, but within Aztec\u2019s own nascent ecosystem. no existing DeFi protocols to compose with yet",
    4: "tagging mechanism requires pre-registering senders. can\u2019t discover notes from unknown senders",
    5: "auditor_elgamal_pubkey exists in mint config, but the entire confidential transfer feature is currently disabled on mainnet pending security audit",
    6: "TFHE is lattice-based (LWE hardness assumption), considered PQ-resistant but less battle-tested than hash-based approaches",
    7: "confidential ERC-20s technically possible via fhEVM, but ecosystem is nascent with limited live deployments",
    8: "ZK rollup architecture is inherently scalable, but currently ~1 TPS in Alpha with 36\u201372s block times. targeting 3\u20134s by end of 2026",
    9: "access-control privacy model (parties see only their part of a transaction). fundamentally different from ZK-based privacy: counterparties and validators see your data",
   10: "DeFi exists within Canton\u2019s own Daml ecosystem. requires Daml, a separate smart contract language. not composable with Ethereum/Starknet DeFi",
   11: "horizontal scaling via network-of-networks. each node only processes its own transactions. claims no upper bound on TPS, but access-control privacy model",
   12: "mainnet since Sep 2024. USDCx with Circle launched Jan 2026. 350+ apps on testnet. but DeFi ecosystem still nascent",
   13: "off-chain execution model enables scalable private computation, but no composability between private applications limits real-world DeFi throughput",
   14: "Trezor supports Orchard shielded transactions. Ledger supports transparent Zcash only",
   15: "PXE client-side proving works but proving times still being optimized. variable depending on circuit complexity",
   16: "can outsource to Prover Market (Proof-as-a-Service). client-side proving possible but resource-intensive",
   17: "Ledger supports Solana but confidential transfer feature is currently disabled on mainnet",
   18: "Orchard proofs ~2\u20135s on modern hardware, but limited to simple shielded transfer logic",
   19: "Bulletproofs+ for amount hiding are formally ZK, but core sender privacy relies on ring signatures which are a different cryptographic mechanism",
   20: "Bulletproofs for range proofs are formally ZK, but scope is limited to amount hiding only",
   21: "always private by default. no shielding needed (instant opt-in), but no way to opt-out to a transparent state",
   22: "1-hour mandatory standby period after shielding before funds can be used privately (Private Proof of Innocence requirement)",
   23: "36\u201372s block times in Alpha. shielding speed will improve as block times target 3\u20134s by late 2026",
   24: "~5s end-to-end from user perspective, but proof runs on wallet operator backend (48-core machine). Zcash Zashi generates proofs directly on mobile in ~5s. for an apples-to-apples comparison, proof generation is not yet on consumer hardware. mobile proving on the roadmap",
}

#                                          0     1     2     3     4     5     6     7     8     9    10    11    12    13    14    15    16
# Features:                            multi  fZK  secr  disc  sdisc noTS    PQ  succ  shld  e2e   scal  prog  DeFi  ecos   EVM    AA    HW
# Groups:                              |---- privacy model ----|  |-- crypto foundation --|  |---- performance & UX ----|  |-- ecosystem & access --|
projects = [
    ("STRK20s (Starknet)", [True, False, True, True, True, True, True, True, True, "partial", True, True, True, True, False, True, False], True, {9: 24}),
    ("Zcash", [False, True, False, False, True, "partial", False, False, False, "partial", False, False, False, True, False, False, "partial"], False, {5: 2, 16: 14, 9: 18}),
    ("Monero", [False, False, False, False, "partial", True, False, False, "partial", False, False, False, False, True, False, False, True], False, {4: 1, 8: 21}),
    ("Aztec", [True, True, False, "partial", True, False, False, True, "partial", "partial", "partial", True, "partial", False, False, True, False], False, {12: 3, 3: 4, 10: 8, 9: 15, 8: 23}),
    ("Aleo", [False, True, False, False, True, False, False, False, False, "partial", "partial", True, False, "partial", False, False, False], False, {13: 12, 10: 13, 9: 16}),
    ("Railgun", [True, True, False, False, True, False, False, False, False, False, False, False, True, True, True, False, False], False, {}),
    ("Solana confidential", [False, False, False, False, "partial", True, False, False, False, False, False, False, False, True, False, False, "partial"], False, {4: 5, 16: 17}),
    ("Canton", [False, False, False, False, True, True, False, False, False, False, "partial", False, "partial", True, False, False, False], False, {12: 10, 10: 11}),
    ("Tornado Cash", [False, True, False, False, False, False, False, False, False, False, False, False, False, False, True, False, False], False, {}),
    ("Zama (FHE)", ["partial", False, False, False, True, True, "partial", False, False, False, False, False, False, False, True, False, False], False, {0: 7, 6: 6}),
]

ws.sheet_properties.tabColor = ORANGE

# Title
ws.merge_cells('A1:S1')
ws['A1'] = "Privacy Protocols: Property Grid"
ws['A1'].font = title_font
ws['A1'].alignment = Alignment(vertical='center')
ws['A1'].fill = PatternFill('solid', fgColor=DARK)
ws.row_dimensions[1].height = 32

ws.merge_cells('A2:S2')
ws['A2'] = "Starknet Ecosystem \u2022 March 2026 \u2022 Verified against official documentation"
ws['A2'].font = subtitle_font
ws['A2'].fill = PatternFill('solid', fgColor=DARK)
ws.row_dimensions[2].height = 18

for col in range(1, len(features)+3):
    ws.cell(row=3, column=col).fill = PatternFill('solid', fgColor=DARK)
ws.row_dimensions[3].height = 6

# Category header row
CR = 4
cat_font = Font(name='Arial', size=8, bold=True, italic=True, color='7A7490')
cat_fill = PatternFill('solid', fgColor='13102A')
cat_align = Alignment(horizontal='center', vertical='center')

# Protocol label cell
c = ws.cell(row=CR, column=1)
c.fill = cat_fill
c.border = thin

# Privacy model: cols 2-6 (5 properties)
ws.merge_cells(start_row=CR, start_column=2, end_row=CR, end_column=6)
c = ws.cell(row=CR, column=2, value="PRIVACY MODEL")
c.font = cat_font
c.fill = cat_fill
c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
c.border = thin
for col in range(3, 7):
    ws.cell(row=CR, column=col).border = thin

# Cryptographic foundation: cols 7-9 (3 properties)
ws.merge_cells(start_row=CR, start_column=7, end_row=CR, end_column=9)
c = ws.cell(row=CR, column=7, value="CRYPTOGRAPHIC FOUNDATION")
c.font = cat_font
c.fill = cat_fill
c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
c.border = thin
for col in range(8, 10):
    ws.cell(row=CR, column=col).border = thin

# Performance & UX: cols 10-13 (4 properties)
ws.merge_cells(start_row=CR, start_column=10, end_row=CR, end_column=13)
c = ws.cell(row=CR, column=10, value="PERFORMANCE & UX")
c.font = cat_font
c.fill = cat_fill
c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
c.border = thin
for col in range(11, 14):
    ws.cell(row=CR, column=col).border = thin

# Ecosystem & access: cols 14-18 (5 properties)
ws.merge_cells(start_row=CR, start_column=14, end_row=CR, end_column=18)
c = ws.cell(row=CR, column=14, value="ECOSYSTEM & ACCESS")
c.font = cat_font
c.fill = cat_fill
c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
c.border = thin
for col in range(15, 19):
    ws.cell(row=CR, column=col).border = thin

# Score cell in category row
c = ws.cell(row=CR, column=len(features)+2)
c.fill = cat_fill
c.border = thin

ws.row_dimensions[CR].height = 18

# Header row
HR = 5
c = ws.cell(row=HR, column=1, value="Protocol")
c.font = Font(name='Arial', size=10, bold=True, color=ORANGE)
c.fill = header_fill
c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
c.border = thin

for i, feat in enumerate(features):
    c = ws.cell(row=HR, column=i+2, value=feat)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = thin

sc = ws.cell(row=HR, column=len(features)+2, value="Score")
sc.font = Font(name='Arial', size=9, bold=True, color=ORANGE)
sc.fill = header_fill
sc.alignment = center
sc.border = thin

ws.row_dimensions[HR].height = 40

# Definition row (short descriptions under each property)
DR = HR + 1
def_font_style = Font(name='Arial', size=7, italic=True, color='7A7490')
def_fill = PatternFill('solid', fgColor='13102A')
def_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

c = ws.cell(row=DR, column=1)
c.fill = def_fill
c.border = thin

for i, sdef in enumerate(short_defs):
    c = ws.cell(row=DR, column=i+2, value=sdef)
    c.font = def_font_style
    c.fill = def_fill
    c.alignment = def_align
    c.border = thin

c = ws.cell(row=DR, column=len(features)+2)
c.fill = def_fill
c.border = thin

ws.row_dimensions[DR].height = 28

ws.column_dimensions['A'].width = 22
for i in range(len(features)):
    ws.column_dimensions[get_column_letter(i+2)].width = 13
ws.column_dimensions[get_column_letter(len(features)+2)].width = 8

# Project rows
for row_idx, (name, ticks, is_starknet, fn_map) in enumerate(projects):
    r = DR + 1 + row_idx
    
    # Count score: True = 1, partial = 0.5, False = 0
    score = sum(1 for t in ticks if t is True) + sum(0.5 for t in ticks if t == "partial")
    score_str = f"{score:g}/17"

    if is_starknet:
        bg = starknet_fill
        nf = starknet_font
        tf_full = Font(name='Arial', size=13, bold=True, color=DARK)
        tf_partial = Font(name='Arial', size=11, bold=True, color='6B4420')
        df = Font(name='Arial', size=10, color='8B5E3C')
        sf = Font(name='Arial', size=11, bold=True, color=DARK)
    else:
        bg = row_fill_even if row_idx % 2 == 0 else row_fill_odd
        nf = name_font
        tf_full = tick_font
        tf_partial = partial_font
        df = dash_font
        sf = Font(name='Arial', size=11, bold=True, color=ORANGE)

    nc = ws.cell(row=r, column=1, value=name)
    nc.font = nf
    nc.alignment = Alignment(vertical='center')
    nc.border = thin
    nc.fill = bg

    for col_idx, has_tick in enumerate(ticks):
        c = ws.cell(row=r, column=col_idx+2)
        if has_tick is True:
            c.value = "\u2713"
            c.font = tf_full
        elif has_tick == "partial":
            fn_num = fn_map.get(col_idx, "")
            c.value = f"(\u2713)\u00B9\u00B2\u00B3\u2074\u2075\u2076\u2077"[0:3]  # placeholder
            # Actually set the proper superscript footnote
            fn_num = fn_map.get(col_idx, "")
            superscripts = {1: "\u00B9", 2: "\u00B2", 3: "\u00B3", 4: "\u2074", 5: "\u2075", 6: "\u2076", 7: "\u2077", 8: "\u2078", 9: "\u2079", 10: "\u00B9\u2070", 11: "\u00B9\u00B9", 12: "\u00B9\u00B2", 13: "\u00B9\u00B3", 14: "\u00B9\u2074", 15: "\u00B9\u2075", 16: "\u00B9\u2076", 17: "\u00B9\u2077", 18: "\u00B9\u2078", 19: "\u00B9\u2079", 20: "\u00B2\u2070", 21: "\u00B2\u00B9", 22: "\u00B2\u00B2", 23: "\u00B2\u00B3", 24: "\u00B2\u2074", 25: "\u00B2\u2075"}
            sup = superscripts.get(fn_num, "")
            c.value = f"(\u2713){sup}"
            c.font = tf_partial
        else:
            c.value = "\u2013"
            c.font = df
        c.alignment = center
        c.border = thin
        c.fill = bg

    # Score
    sc = ws.cell(row=r, column=len(features)+2, value=score_str)
    sc.font = sf
    sc.alignment = center
    sc.border = thin
    sc.fill = bg

    ws.row_dimensions[r].height = 30

# Legend
lr = DR + len(projects) + 2
ws.cell(row=lr, column=1, value="\u2713 = full capability").font = Font(name='Arial', size=9, color=ORANGE)
ws.cell(row=lr, column=1).fill = PatternFill('solid', fgColor=DARK)
ws.cell(row=lr, column=3, value="(\u2713) = capability with caveats (see footnotes)").font = Font(name='Arial', size=9, color='C4883D')
ws.cell(row=lr, column=3).fill = PatternFill('solid', fgColor=DARK)
ws.cell(row=lr, column=7, value="\u2013 = absent or not applicable").font = Font(name='Arial', size=9, color='3D3658')
ws.cell(row=lr, column=7).fill = PatternFill('solid', fgColor=DARK)

for col in range(1, len(features)+3):
    ws.cell(row=lr, column=col).fill = PatternFill('solid', fgColor=DARK)

# Footnotes on main sheet
fn_start = lr + 2
ws.cell(row=fn_start, column=1, value="Footnotes").font = Font(name='Arial', size=10, bold=True, color=ORANGE)
ws.cell(row=fn_start, column=1).fill = PatternFill('solid', fgColor=DARK)
for col in range(1, len(features)+3):
    ws.cell(row=fn_start, column=col).fill = PatternFill('solid', fgColor=DARK)

superscripts_display = {1: "\u00B9", 2: "\u00B2", 3: "\u00B3", 4: "\u2074", 5: "\u2075", 6: "\u2076", 7: "\u2077", 8: "\u2078", 9: "\u2079", 10: "\u00B9\u2070", 11: "\u00B9\u00B9", 12: "\u00B9\u00B2", 13: "\u00B9\u00B3", 14: "\u00B9\u2074", 15: "\u00B9\u2075", 16: "\u00B9\u2076", 17: "\u00B9\u2077", 18: "\u00B9\u2078", 19: "\u00B9\u2079", 20: "\u00B2\u2070", 21: "\u00B2\u00B9", 22: "\u00B2\u00B2", 23: "\u00B2\u00B3", 24: "\u00B2\u2074", 25: "\u00B2\u2075"}

for fn_num in sorted(footnotes.keys()):
    row = fn_start + fn_num
    sup = superscripts_display[fn_num]
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(features)+2)
    c = ws.cell(row=row, column=1, value=f"{sup}  {footnotes[fn_num]}")
    c.font = Font(name='Arial', size=8, italic=True, color='9990B0')
    c.fill = PatternFill('solid', fgColor=DARK)
    c.alignment = Alignment(wrap_text=True, vertical='top')
    for col in range(1, len(features)+3):
        ws.cell(row=row, column=col).fill = PatternFill('solid', fgColor=DARK)

# Fill remaining rows dark
for row in range(fn_start + 20, fn_start + 24):
    for col in range(1, len(features)+3):
        ws.cell(row=row, column=col).fill = PatternFill('solid', fgColor=DARK)

# === Definitions & rationale sheet ===
ws2 = wb.create_sheet("Definitions & rationale")
ws2.sheet_properties.tabColor = ORANGE

ws2['A1'] = "Property"
ws2['B1'] = "Definition & Rationale"
ws2['A1'].font = Font(name='Arial', size=10, bold=True, color=ORANGE)
ws2['B1'].font = Font(name='Arial', size=10, bold=True, color=ORANGE)
ws2['A1'].fill = PatternFill('solid', fgColor=DARK)
ws2['B1'].fill = PatternFill('solid', fgColor=DARK)
ws2.column_dimensions['A'].width = 28
ws2.column_dimensions['B'].width = 120

definitions = [
    # Privacy model
    ("Multi-asset privacy pool", "single pool supports multiple token types. observers can\u2019t tell which token a note belongs to. STRK20s, Aztec, and Railgun have full multi-asset pools. Zama supports confidential ERC-20s via fhEVM but ecosystem is nascent. Aleo is asset-agnostic but each program has isolated state with no composability between private applications, so no unified pool. Zcash/Monero are single-asset. Tornado Cash had fragmented pools per token per amount. Canton uses access-control privacy, not a shielded pool model."),
    ("Formally ZK proofs", "the proof system has the mathematical zero-knowledge property: the proof itself guarantees that no information about the secret data leaks to the verifier. This is distinct from privacy-as-UX. Many ZK rollups (including Starknet today) use validity proofs that prove correct execution but do not have the formal ZK property, meaning the proof could theoretically leak information about the underlying computation. Aztec\u2019s CEO calls their architecture a \u2018zk-zk rollup\u2019: inner SNARKs (per-transaction) have the formal ZK property, outer SNARKs (rollup) also have it. Zcash (Groth16/Halo 2), Railgun (Groth16), Aleo (Varuna), and Tornado Cash (Groth16) all use proof systems with the formal ZK guarantee. STRK20s uses STARKs that are not formally ZK yet (Adrien confirmed). Monero\u2019s Bulletproofs+ are formally ZK but only cover amount hiding; core sender privacy comes from ring signatures. Solana\u2019s Bulletproofs are formally ZK but limited to range proofs. Zama uses FHE (different model). Canton uses access control (no ZK proofs)."),
    ("Secretless", "users handle a single signing key (or none, via AA). no per-note secrets to manage or lose. STRK20s derives all keys from the Starknet account key via account abstraction. Aztec requires 3 separate key pairs. Aleo requires private key + view key. Railgun requires spending + viewing keys. Monero requires spend + view keys. Canton uses party keys managed via validator nodes. Others also require dedicated privacy key management."),
    ("Efficient note discovery", "recipients discover funds in time proportional to their own activity. STRK20s uses on-chain indexed storage with shared secrets. Aztec uses a tagging mechanism but requires pre-registering senders. Zcash requires scanning all blocks and decrypting every transaction. Railgun and Aleo use standard record/UTXO scanning. Canton uses a ledger model, not applicable."),
    ("Selective disclosure", "users can optionally reveal transaction history to a chosen third party (auditor, accountant, regulator, DAO) without affecting other users\u2019 privacy. whether you use this for regulatory compliance, tax reporting, or governance is up to you. STRK20s has third-party viewing keys. Zcash has full incoming/outgoing viewing keys. Aztec has incoming + outgoing viewing key pairs. Aleo has view keys plus Selective Disclosure and zPass for identity proofs. Railgun has scoped viewing keys (by block range) plus PPOI plus Koinly tax exports. Canton has sub-transaction privacy with full auditability built into Daml contracts. Solana has auditor_elgamal_pubkey in mint config but feature is disabled. Monero has viewing keys but incoming only, all-or-nothing. Zama has programmable ACL contracts. Tornado Cash had a voluntary report tool only."),
    # Cryptographic foundation
    ("No trusted setup", "proof system requires no trusted setup ceremony. STARKs (Starknet) require none. Zcash Orchard uses Halo 2 (no setup) but Sapling pool still has one. Monero\u2019s Bulletproofs+ and Solana\u2019s ElGamal/Bulletproofs require none. Zama\u2019s TFHE requires none. Canton doesn\u2019t use ZK proofs, no ceremony needed. Aleo uses Varuna (Marlin-based universal SNARK) requiring a universal SRS. Aztec\u2019s PLONK uses a universal updatable SRS (still a form of setup). Railgun and Tornado Cash use Groth16 requiring per-circuit ceremonies."),
    ("Post-quantum resistant", "underlying cryptography resists quantum computing attacks. STARKs (hash-based) are post-quantum by design. Zama\u2019s TFHE is lattice-based (LWE), considered PQ-resistant but less battle-tested. All SNARK/elliptic curve systems (Zcash, Aztec, Aleo, Railgun, Monero, Solana, Canton) are vulnerable to quantum attacks."),
    ("Succinct verifiability", "anyone can verify chain integrity quickly and trustlessly on a small device. Enables rich client-side privacy proofs. ZK rollups (Starknet, Aztec) have this. L1s like Solana, Zcash, Monero, Aleo, and Canton do not. Aleo posts per-transaction proofs but doesn\u2019t provide chain-level succinct state verification. Railgun runs on L1s without succinct verification."),
    # Performance & UX
    ("Fast shield/unshield", "how quickly a user can move between public and private states. measures the end-to-end time from initiating a shield (opt-in to privacy) or unshield (opt-out) to having usable funds. STRK20s: sub-2s block confirmation + sub-5s proof generation = seconds to shield or unshield. fastest of any protocol. Zcash: ~75s block time is the bottleneck. proof generation 2\u20135s but must wait for block confirmation. Monero: always private by default (instant opt-in), but no way to opt-out to a transparent state. Aztec: 36\u201372s block times in Alpha (targeting 3\u20134s by late 2026). Railgun: Ethereum block time (~12s) for the on-chain tx, but a mandatory 1-hour standby period after shielding before funds can be used (Private Proof of Innocence requirement). Tornado Cash: deposit was one Ethereum block, but real privacy required waiting hours for anonymity set to build. Aleo: off-chain proving but no concrete shielding speed benchmarks available. Solana: feature disabled on mainnet. Canton: access control model, no shielding concept. Zama: FHE operations are orders of magnitude slower than ZK."),
    ("End-to-end transaction time", "end-to-end transaction time from user action to confirmed private transaction. the benchmark is Zcash Zashi: ~5 seconds on mobile, proof generated directly on the device. STRK20s: ~5 seconds end-to-end from the user\u2019s perspective, but proof generation runs on the wallet operator backend (48-core machine), not on the user\u2019s device. mobile proving is on the roadmap. for an apples-to-apples comparison, Zcash would also be faster if it ran proofs server-side. Aztec: PXE client-side proving works but proving times are still being optimized in Alpha. Aleo: client-side proving possible but resource-intensive; outsourceable to Prover Market. Railgun: Groth16 proving can be slow for complex circuits. Monero: ring signature computation is lightweight but not ZK. Zama: FHE is orders of magnitude slower than ZK. Canton, Solana: no ZK proving involved."),
    ("Scalable private transactions", "high throughput and low cost for private operations specifically. Starknet: ~1,000 TPS today, targeting 10k+. sub-2s confirmation, $0.002 avg fee. Canton: claims no upper bound on TPS via horizontal scaling but access-control privacy model. Aleo: off-chain execution model is architecturally scalable but no composability between private apps limits real-world DeFi throughput. Zcash: PoW L1, ~75s blocks. Monero: PoW L1, 2-min blocks. Aztec: ~1 TPS in Alpha, 36\u201372s blocks. Railgun: inherits L1 gas costs. Solana: fast but privacy feature disabled. Zama: FHE is orders of magnitude slower. Tornado Cash: L1 gas costs."),
    ("Programmable privacy", "ability to prove complex logic (multi-calls, full smart contract execution, atomic DeFi operations) privately, verified cheaply on-chain. STRK20s: Cairo + Stwo enables proving entire Starknet block logic privately. atomic private AMM swaps (withdraw + swap + redeposit in one tx), account abstraction, session keys all work within the privacy layer. the highest ceiling of any privacy protocol. Aztec: Noir + PXE executes private functions in browser with ZK proofs. Aleo: Leo programs execute off-chain with ZK proofs, but no composability between private programs limits multi-step atomic operations. Zcash: limited to simple shielded transfers. Tornado Cash: \u201CI have enough funds\u201D was the ceiling. Railgun: bounded proof complexity. Solana: not succinctly verifiable. Zama: server-side FHE on coprocessors. Canton: no ZK proving, privacy via access control."),
    # Ecosystem & access
    ("DeFi composability", "existing DeFi contracts require zero custom integration for private transactions. STRK20s and Railgun work with existing live protocols. Aztec and Canton have DeFi within their own ecosystems. Aleo has no composability between private applications (Equilibrium Labs: \u2018one application can\u2019t use the private state from another application as its input\u2019). Zama and Solana require DeFi protocols to rewrite for encrypted types."),
    ("Existing ecosystem", "live DeFi protocols, real liquidity, real users already present before privacy is added. Starknet, Zcash, Monero, Railgun (on Ethereum/Polygon/Arbitrum/BSC), Solana, and Canton ($4T annual tokenized volume, Goldman Sachs, HSBC, DTCC, Euroclear) all have established ecosystems. Aleo launched mainnet Sep 2024 and USDCx with Circle Jan 2026, but DeFi is still nascent. Aztec is in Alpha with no live DeFi. Zama\u2019s ecosystem is nascent. Tornado Cash is OFAC-sanctioned."),
    ("EVM compatible", "privacy protocol runs on or integrates natively with EVM chains. Railgun operates as a smart contract on Ethereum, Polygon, Arbitrum, BSC. Tornado Cash was Ethereum-native. Zama\u2019s fhEVM is an EVM-compatible confidential computing layer. STRK20s runs on Starknet (Cairo VM, not EVM). Aztec uses Noir. Aleo uses Leo. Zcash, Monero, Solana, Canton are all non-EVM chains."),
    ("Account abstraction", "privacy works with multisig, session keys, social recovery, smart accounts at the protocol level. STRK20s and Aztec both have native protocol-level AA. Monero has experimental CLI-only multisig via key splitting (not protocol-level). Aleo uses standard private key + view key + address model. Canton uses a party/validator model, not crypto-style AA. Others require standard key management."),
    ("Hardware wallet support", "users can secure private keys and sign privacy transactions with a hardware wallet. Starknet has Ledger support for standard transactions (via Ledger Live, Braavos, Argent X, and Ready multisig), but STRK20s privacy transactions are not yet available on hardware wallets. Monero: Ledger supports full shielded transactions. Zcash: Trezor supports Orchard shielded transactions, but Ledger supports transparent Zcash only. Solana: Ledger supports Solana, but confidential transfer feature is currently disabled on mainnet. Aztec, Aleo, Railgun, Zama, Canton, Tornado Cash: no hardware wallet integration for privacy features."),
]

categories = [
    (0, "PRIVACY MODEL"),
    (5, "CRYPTOGRAPHIC FOUNDATION"),
    (8, "PERFORMANCE & UX"),
    (12, "ECOSYSTEM & ACCESS"),
]
cat_indices = {idx: name for idx, name in categories}

r = 2
for i, (term, defn) in enumerate(definitions):
    if i in cat_indices:
        ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        c = ws2.cell(row=r, column=1, value=cat_indices[i])
        c.font = Font(name='Arial', size=9, bold=True, italic=True, color=ORANGE)
        c.fill = PatternFill('solid', fgColor=DARK)
        c.alignment = Alignment(vertical='center')
        ws2.cell(row=r, column=2).fill = PatternFill('solid', fgColor=DARK)
        ws2.row_dimensions[r].height = 22
        r += 1
    a = ws2.cell(row=r, column=1, value=term)
    a.font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
    a.fill = PatternFill('solid', fgColor=DARK_LIGHT)
    b = ws2.cell(row=r, column=2, value=defn)
    b.font = Font(name='Arial', size=9, color='B0ACC0')
    b.fill = PatternFill('solid', fgColor=DARK_LIGHT)
    b.alignment = Alignment(wrap_text=True)
    r += 1

ws.freeze_panes = 'B7'

# === Marketing summary tab ===
ws3 = wb.create_sheet("At a glance")
ws3.sheet_properties.tabColor = ORANGE

mkt_features = [
    "Private by\ndesign",
    "Composable\nfunds",
    "Fast to\nuse",
    "Selective\ndisclosure",
    "Future-proof\nsecurity",
    "Wallet\nready",
]

mkt_projects = [
    ("STRK20s (Starknet)", [True, True, True, True, True, "partial"], True, {5: 1}),
    ("Zcash", ["partial", False, False, True, False, "partial"], False, {0: 2, 5: 3}),
    ("Monero", ["partial", False, "partial", "partial", False, "partial"], False, {0: 4, 2: 17, 3: 5, 5: 6}),
    ("Aztec", [True, "partial", False, True, "partial", "partial"], False, {1: 7, 4: 8, 5: 9}),
    ("Aleo", ["partial", False, False, True, False, False], False, {0: 10}),
    ("Railgun", ["partial", True, False, True, False, False], False, {0: 11}),
    ("Solana confidential", [False, False, False, "partial", False, False], False, {3: 12}),
    ("Canton", [False, "partial", False, True, False, False], False, {1: 13}),
    ("Tornado Cash", ["partial", False, False, False, False, False], False, {0: 15}),
    ("Zama (FHE)", ["partial", False, False, True, "partial", False], False, {0: 16, 4: 14}),
]

mkt_footnotes = {
    1: "Starknet has full AA and wallet integration (Braavos, Argent), but STRK20s privacy transactions are not yet available on hardware wallets",
    2: "formally ZK proofs, but single-asset only (ZEC), no multi-asset pool, poor note discovery UX",
    3: "Trezor supports Orchard shielded transactions. Ledger only transparent",
    4: "always-on privacy with strongest anonymity set, but no multi-asset pool, no formal ZK proofs, no secretless design",
    5: "incoming viewing keys only. can\u2019t reliably verify outgoing or balance",
    6: "Ledger supports full shielded transactions, but no account abstraction",
    7: "smart contracts exist but zero existing DeFi or liquidity. cold start problem",
    8: "PLONK uses universal SRS (still a setup). no post-quantum. but has succinct verifiability",
    9: "native AA, but no hardware wallet support for privacy transactions",
    10: "formally ZK but no multi-asset pool, no composability between private apps, no note discovery",
    11: "formally ZK and multi-asset, but privacy limited by pool size per token. no client-side proving of complex logic",
    12: "auditor key exists in config but entire feature is disabled on mainnet",
    13: "DeFi within own Daml ecosystem only. requires separate smart contract language",
    14: "TFHE is lattice-based (PQ-resistant) and no trusted setup, but no succinct verifiability",
    15: "formally ZK (Groth16), but fragmented pools (fixed denominations), timing analysis could de-anonymize users, and protocol is OFAC-sanctioned/defunct",
    16: "FHE genuinely encrypts data during computation, but server-side model (not client-side ZK proofs). different trust model. ecosystem nascent",
    17: "privacy is always on so there\u2019s no shielding step (zero friction to opt in). ~2 min block time for confirmations. seamless UX but slower than Starknet",
}

mkt_superscripts = {1: "\u00B9", 2: "\u00B2", 3: "\u00B3", 4: "\u2074", 5: "\u2075", 6: "\u2076", 7: "\u2077", 8: "\u2078", 9: "\u2079", 10: "\u00B9\u2070", 11: "\u00B9\u00B9", 12: "\u00B9\u00B2", 13: "\u00B9\u00B3", 14: "\u00B9\u2074", 15: "\u00B9\u2075", 16: "\u00B9\u2076", 17: "\u00B9\u2077"}

NCOLS = len(mkt_features) + 1

# Solid visible border for Google Sheets (slightly brighter than the dark bg so it shows)
mkt_border = Border(
    left=Side(style='thin', color='3D3658'),
    right=Side(style='thin', color='3D3658'),
    top=Side(style='thin', color='3D3658'),
    bottom=Side(style='thin', color='3D3658')
)

# Fill ALL cells in the used range with dark background to avoid white gaps
for r in range(1, 35 + len(mkt_projects) + len(mkt_footnotes)):
    for c in range(1, NCOLS + 3):
        ws3.cell(row=r, column=c).fill = PatternFill('solid', fgColor=DARK)

# Title
ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOLS)
ws3['A1'] = "Privacy Protocols: At a Glance"
ws3['A1'].font = title_font
ws3['A1'].alignment = Alignment(vertical='center')
ws3.row_dimensions[1].height = 32

ws3.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NCOLS)
ws3['A2'] = "what matters to builders and users \u2022 March 2026"
ws3['A2'].font = subtitle_font
ws3.row_dimensions[2].height = 18

ws3.row_dimensions[3].height = 6

# Header
MHR = 4
c = ws3.cell(row=MHR, column=1, value="Protocol")
c.font = Font(name='Arial', size=10, bold=True, color=ORANGE)
c.fill = header_fill
c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
c.border = mkt_border

for i, feat in enumerate(mkt_features):
    c = ws3.cell(row=MHR, column=i+2, value=feat)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = mkt_border

ws3.row_dimensions[MHR].height = 40

ws3.column_dimensions['A'].width = 22
for i in range(len(mkt_features)):
    ws3.column_dimensions[get_column_letter(i+2)].width = 16

# Description row
desc_row = MHR + 1
descs = [
    "is my stuff\nactually private?",
    "can I use\nprivate funds freely?",
    "is the\nexperience fast?",
    "can institutions\nuse this legally?",
    "will this still be\nsecure in 5 years?",
    "does my\nwallet work?",
]
c = ws3.cell(row=desc_row, column=1)
c.fill = PatternFill('solid', fgColor='13102A')
c.border = mkt_border
for i, d in enumerate(descs):
    c = ws3.cell(row=desc_row, column=i+2, value=d)
    c.font = Font(name='Arial', size=8, italic=True, color='7A7490')
    c.fill = PatternFill('solid', fgColor='13102A')
    c.alignment = center
    c.border = mkt_border
ws3.row_dimensions[desc_row].height = 30

# Project rows
for row_idx, (name, ticks, is_starknet, fn_map) in enumerate(mkt_projects):
    r = desc_row + 1 + row_idx

    if is_starknet:
        bg = starknet_fill
        nf = starknet_font
        tf_full = Font(name='Arial', size=13, bold=True, color=DARK)
        tf_partial = Font(name='Arial', size=11, bold=True, color='6B4420')
        df = Font(name='Arial', size=10, color='8B5E3C')
    else:
        bg = row_fill_even if row_idx % 2 == 0 else row_fill_odd
        nf = name_font
        tf_full = tick_font
        tf_partial = partial_font
        df = dash_font

    nc = ws3.cell(row=r, column=1, value=name)
    nc.font = nf
    nc.alignment = Alignment(vertical='center')
    nc.border = mkt_border
    nc.fill = bg

    for col_idx, has_tick in enumerate(ticks):
        c = ws3.cell(row=r, column=col_idx+2)
        if has_tick is True:
            c.value = "\u2713"
            c.font = tf_full
        elif has_tick == "partial":
            fn_num = fn_map.get(col_idx, "")
            sup = mkt_superscripts.get(fn_num, "") if fn_num else ""
            c.value = f"(\u2713){sup}"
            c.font = tf_partial
        else:
            c.value = "\u2013"
            c.font = df
        c.alignment = center
        c.border = mkt_border
        c.fill = bg

    ws3.row_dimensions[r].height = 30

# Legend
mlr = desc_row + 1 + len(mkt_projects) + 1
ws3.cell(row=mlr, column=1, value="\u2713 = yes").font = Font(name='Arial', size=9, color=ORANGE)
ws3.cell(row=mlr, column=3, value="(\u2713) = partially (see footnotes)").font = Font(name='Arial', size=9, color='C4883D')
ws3.cell(row=mlr, column=6, value="\u2013 = no").font = Font(name='Arial', size=9, color='3D3658')

# Footnotes
mfn_start = mlr + 2
ws3.cell(row=mfn_start, column=1, value="Footnotes").font = Font(name='Arial', size=10, bold=True, color=ORANGE)

for fn_num in sorted(mkt_footnotes.keys()):
    row = mfn_start + fn_num
    sup = mkt_superscripts[fn_num]
    ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
    c = ws3.cell(row=row, column=1, value=f"{sup}  {mkt_footnotes[fn_num]}")
    c.font = Font(name='Arial', size=8, italic=True, color='9990B0')
    c.alignment = Alignment(wrap_text=True, vertical='top')

ws3.freeze_panes = 'B6'

# === Definitions & rationale (at a glance) tab ===
ws4 = wb.create_sheet("Definitions & rationale 2")
ws4.sheet_properties.tabColor = ORANGE

# Pre-fill dark background
for r in range(1, 20):
    for c in range(1, 6):
        ws4.cell(row=r, column=c).fill = PatternFill('solid', fgColor=DARK)

# Title
ws4.merge_cells('A1:E1')
ws4['A1'] = "At a Glance: Definitions & Rationale"
ws4['A1'].font = title_font
ws4['A1'].alignment = Alignment(vertical='center')
ws4.row_dimensions[1].height = 32

ws4.merge_cells('A2:E2')
ws4['A2'] = "how the 6 summary properties map to the full 17-property grid \u2022 March 2026"
ws4['A2'].font = subtitle_font
ws4.row_dimensions[2].height = 18

ws4.row_dimensions[3].height = 6

# Header
dr_header = 4
ws4.column_dimensions['A'].width = 20
ws4.column_dimensions['B'].width = 60
ws4.column_dimensions['C'].width = 50

for col_label, col_num in [("Property", 1), ("What it means", 2), ("Based on (from full Property Grid)", 3)]:
    c = ws4.cell(row=dr_header, column=col_num, value=col_label)
    c.font = Font(name='Arial', size=9, bold=True, color=ORANGE)
    c.fill = PatternFill('solid', fgColor=DARK_LIGHT)
    c.border = mkt_border
ws4.row_dimensions[dr_header].height = 24

mkt_definitions = [
    ("Private by design",
     "your balances, transfers, and identity are cryptographically hidden. the proof system itself guarantees no data leaks. you don\u2019t manage separate secrets or scan the entire chain to find your funds.",
     "multi-asset privacy pool + formally ZK proofs + secretless + efficient note discovery + programmable privacy"),
    ("Composable funds",
     "your private funds can interact with live DeFi protocols (swap, lend, stake) without leaving the privacy layer or migrating to a new chain.",
     "DeFi composability + existing ecosystem + EVM compatible"),
    ("Fast to use",
     "shielding (going private), unshielding (going public), and private transactions all happen in seconds. no waiting periods, no slow proof generation.",
     "fast shield/unshield + end-to-end transaction time + scalable private transactions"),
    ("Selective disclosure",
     "users can optionally reveal transaction history to a chosen third party (auditor, accountant, regulator, DAO) without affecting anyone else\u2019s privacy.",
     "selective disclosure"),
    ("Future-proof security",
     "the cryptographic foundation will still be secure when quantum computers arrive. no trusted setup ceremony that could be compromised. anyone can verify the chain on a small device.",
     "no trusted setup + post-quantum proof system + succinct verifiability"),
    ("Wallet ready",
     "privacy works with your existing wallet setup: multisig, social recovery, session keys. hardware wallet signing is supported for the underlying chain.",
     "account abstraction + hardware wallet support"),
]

for i, (prop, meaning, based_on) in enumerate(mkt_definitions):
    r = dr_header + 1 + i
    bg_fill = PatternFill('solid', fgColor='151225' if i % 2 == 0 else DARK_LIGHT)

    a = ws4.cell(row=r, column=1, value=prop)
    a.font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
    a.fill = bg_fill
    a.border = mkt_border
    a.alignment = Alignment(vertical='top')

    b = ws4.cell(row=r, column=2, value=meaning)
    b.font = Font(name='Arial', size=9, color='B0ACC0')
    b.fill = bg_fill
    b.border = mkt_border
    b.alignment = Alignment(wrap_text=True, vertical='top')

    d = ws4.cell(row=r, column=3, value=based_on)
    d.font = Font(name='Arial', size=8, italic=True, color='7A7490')
    d.fill = bg_fill
    d.border = mkt_border
    d.alignment = Alignment(wrap_text=True, vertical='top')

    ws4.row_dimensions[r].height = 50

# Fill remaining rows dark
for r in range(dr_header + 1 + len(mkt_definitions), dr_header + 1 + len(mkt_definitions) + 4):
    for c in range(1, 4):
        ws4.cell(row=r, column=c).fill = PatternFill('solid', fgColor=DARK)

wb.save('/home/claude/privacy_feature_grid_final.xlsx')
print("Done")
